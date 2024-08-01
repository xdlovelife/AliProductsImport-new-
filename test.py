import logging
import time
import tkinter as tk
from tkinter import filedialog
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException, StaleElementReferenceException
from openpyxl import load_workbook
from contextlib import contextmanager

# 设置日志记录器的配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s || %(message)s')
logger = logging.getLogger(__name__)

chrome_driver_path = 'D:\\chromedriver-win64\\chromedriver.exe'
chromium_binary_path = 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'
user_data_dir = 'C:\\Users\\Administrator\\AppData\\Local\\Google\\Chrome\\User Data'

options = Options()
options.binary_location = chromium_binary_path
options.add_argument(f'--user-data-dir={user_data_dir}')


@contextmanager
def open_browser():
    attempts = 3
    driver = None
    for attempt in range(1, attempts + 1):
        try:
            driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)
            logger.info("Chrome WebDriver启动成功。")
            yield driver
            return
        except Exception as e:
            logger.error(f"第 {attempt} 次尝试启动Chrome失败: {str(e)}")
            if attempt == attempts:
                logger.error("达到最大重试次数。退出程序。")
                raise
            time.sleep(3)
        finally:
            if driver:
                driver.quit()


def open_alibaba(driver, selected_categories, sheet_names):
    try:
        if driver:
            url = "https://www.alibaba.com/"
            logger.info(f"访问页面: {url}")
            driver.get(url)

            search_bar = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'fy23-icbu-search-bar-inner'))
            )

            for category in selected_categories:
                try:
                    process_link(driver, "https://www.alibaba.com/", category, sheet_names)
                except Exception as e:
                    logger.error(f"处理类别 '{category}' 出错: {e}")

            driver.quit()

    except NoSuchElementException as e:
        logger.error(f"未找到元素：{e}")
    except TimeoutException as e:
        logger.error(f"超时等待元素加载：{e}")
    except Exception as e:
        logger.error(f"发生异常: {str(e)}")


def process_link(driver, link, category, sheet_names):
    try:
        logger.info(f"处理分类: {category}")
        logger.info(f"处理链接: {link}")
        driver.get(link)
        driver.switch_to.window(driver.window_handles[0])

        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input.search-bar-input.util-ellipsis'))
        )
        search_input.clear()
        search_input.send_keys(category)

        search_button = driver.find_element(By.CSS_SELECTOR, 'button.fy23-icbu-search-bar-inner-button')
        search_button.click()

        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, "organic-list"))
        )

        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        product_list = driver.find_elements(By.CLASS_NAME, "fy23-search-card")
        num_products = len(product_list)
        logger.info(f"共抓取到的产品数量：{num_products}")

        success_count = 0
        for product in product_list:
            start_time = time.time()

            try:
                product_title = product.find_element(By.CLASS_NAME, "search-card-e-title")
                logger.info(f"当前产品标题: {product_title.text}")

                scroll_to_element(driver, product_title)
                time.sleep(1)

                product_link = product.find_element(By.TAG_NAME, "a").get_attribute("href")
                driver.execute_script(f"window.open('{product_link}')")

                success_count = handle_product_detail(driver, category, success_count, sheet_names)

                time.sleep(1)

                end_time = time.time()
                processing_time = end_time - start_time
                logger.info(f"处理时间：{processing_time:.2f} 秒")

                screen_width = get_screen_width()
                logger.info("=" * (screen_width // 10))

                success_count += 1
            except NoSuchElementException as e:
                logger.error(f"未找到产品标题或链接: {e}")
            except Exception as e:
                logger.error(f"处理产品时发生错误: {e}")

        logger.info(f"成功处理的产品数量: {success_count}")
        return success_count

    except Exception as e:
        logger.error(f"处理链接时发生错误: {e}")
        return 0


def handle_product_detail(driver, category, success_count, sheet_names):
    try:
        original_window = driver.current_window_handle
        handles = driver.window_handles

        new_window = None
        for window_handle in handles:
            if window_handle != original_window:
                new_window = window_handle
                break

        if new_window:
            driver.switch_to.window(new_window)
            product_title = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.TAG_NAME, "h1")))
            success_count = handle_product_actions(driver, category, success_count, sheet_names)
            time.sleep(1)
            driver.switch_to.window(original_window)
        return success_count
    except NoSuchWindowException as e:
        logger.error(f"浏览器窗口丢失：{e}")
        close_tab(driver, new_window)
        return success_count
    except TimeoutException as e:
        logger.error(f"超时等待产品详情页加载：{e}")
        close_tab(driver, new_window)
        return success_count
    except Exception as e:
        logger.error(f"处理产品详情页时发生错误: {e}")
        close_tab(driver, new_window)
        return success_count


def handle_product_actions(browser, category, success_count, sheet_names):
    logger.info(f"处理产品详情页操作: {category}, {sheet_names}!!!")
    try:
        add_btn_con = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="addBtnCon"]')))
        add_btn_con.click()
        logging.info("点击了按钮//*[@id='addBtnCon']")

        try:
            element = WebDriverWait(browser, 20).until(
                EC.presence_of_element_located((By.XPATH, '//span[@class="inactive" and text()="Draft"]'))
            )
            logging.info("成功加载 Draft 元素")
            actions = ActionChains(browser)
            actions.move_to_element(element).perform()
            element.click()
            logging.info("成功点击 Draft 元素")
            time.sleep(2)
        except Exception as e:
            logging.error(f"等待和点击 Draft 元素时出现错误：{e}")
            close_current_tab(browser)
            return success_count

        time.sleep(3)

        success_message = None
        try:
            success_message = browser.find_element(By.XPATH,
                                                   '//div[@class="textcontainer centeralign home-content "]/p[1]')
            if success_message.text == "This product is already in your store, what would you like to do?":
                logging.info("产品已存在，不再处理当前产品")
            browser.close()
            return success_count
        except NoSuchElementException:
            pass

        time.sleep(2)

        select_button = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="ms-choice"]'))
        )
        logging.info("等待并点击选择按钮")
        select_button.click()

        dropdown = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "ms-drop"))
        )

        search_input = dropdown.find_element(By.CSS_SELECTOR, ".ms-search input[type='text']")
        search_input.clear()
        search_input.send_keys(sheet_names)
        logging.info(f"输入关键词: {sheet_names}")

        fetch_dropdown_options(browser, sheet_names)
        time.sleep(3)

        try:
            description_tab_button = browser.find_element(By.XPATH, '//*[@id="description_tab_button"]')
            description_tab_button.click()
            logging.info("点击了 description_tab_button 按钮")
            time.sleep(3)

            variants_button = browser.find_element(By.CSS_SELECTOR,
                                                   'button.accordion-tab[data-actab-group="0"][data-actab-id="2"]')
            variants_button.click()
            logging.info("点击了 Variants 按钮")

            all_variants_radio = browser.find_element(By.ID, 'all_variants')
            all_variants_radio.click()
            logging.info("选择 Import all variants automatically 单选框")

            time.sleep(3)

            price_switch_radio = browser.find_element(By.ID, 'price_switch')
            price_switch_radio.click()
            logging.info("选择 Select product price 单选框")

            time.sleep(3)

            save_btn = WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[@class="save"]'))
            )
            save_btn.click()
            logging.info("点击了保存按钮")
            success_count += 1

        except Exception as e:
            logging.error(f"在产品操作中出现错误：{e}")
            close_current_tab(browser)
            return success_count

        time.sleep(3)
        close_current_tab(browser)

    except Exception as e:
        logging.error(f"处理产品操作时发生错误：{e}")

    return success_count


def fetch_dropdown_options(browser, sheet_names):
    dropdown = WebDriverWait(browser, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, "ms-drop"))
    )

    options = dropdown.find_elements(By.TAG_NAME, "li")

    for option in options:
        # 检查 sheet_names 是否为列表
        if isinstance(sheet_names, list):
            for name in sheet_names:
                if name in option.text:
                    option.click()
                    logging.info(f"点击选择的选项: {option.text}")
                    return
        else:
            if sheet_names in option.text:
                option.click()
                logging.info(f"点击选择的选项: {option.text}")
                return


def close_current_tab(browser):
    try:
        browser.close()
    except Exception as e:
        logging.error(f"关闭当前标签页时发生错误: {e}")

def close_tab(driver, window_handle):
    try:
        if window_handle:
            driver.switch_to.window(window_handle)
            driver.close()
            logger.info("关闭出错的产品详情页标签页")
    except Exception as e:
        logger.error(f"关闭标签页时发生错误: {e}")

def get_screen_width():
    return 150

def scroll_to_element(browser, element):
    try:
        WebDriverWait(browser, 60).until(EC.visibility_of(element))
        browser.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        logging.info(f"滚动到元素: {element.text}")
    except Exception as e:
        logging.error(f"滚动到元素时出错: {e}")

def main():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        logger.error("未选择文件。程序退出。")
        return

    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames

        selected_categories = []
        for sheet_name in sheet_names:
            if sheet_name != 'Sheet1':  # 排除Sheet1
                selected_categories.append(sheet_name)

        logger.info(f"选中的类别：{selected_categories}")
        logger.info(f"工作表名：{sheet_names}")

        with open_browser() as driver:
            open_alibaba(driver, selected_categories, sheet_names)

    except Exception as e:
        logger.error(f"处理 Excel 文件时发生错误: {e}")


if __name__ == '__main__':
    main()
